"""Customer connector self-setup via a private, tokenised link.

Flow: the operator picks connectors for a company and sends an access request.
The customer's IT admin opens a private link, enters the read-only credentials
themselves (over HTTPS, stored encrypted — never by email), and the app runs a
live read-only check to confirm it worked. This module holds the token
lifecycle, the instruction workbook, and the connection test.
"""
from __future__ import annotations

import hashlib
import io
import secrets
from datetime import datetime, timedelta, timezone

from sqlalchemy import select
from sqlalchemy.orm import Session

from ..models import AccessRequest, Company, Connector

# provider -> scanner module (same set the assessment uses)
_SCANNERS = {
    "aws": "app.domain.scanners.aws",
    "azure": "app.domain.scanners.azure",
    "intune": "app.domain.scanners.intune",
    "gcp": "app.domain.scanners.gcp",
    "adgpo": "app.domain.scanners.adgpo",
    "firewall": "app.domain.scanners.firewall",
}

# Step-by-step the customer's IT admin performs, per connector. Read-only throughout.
SETUP_STEPS = {
    "aws": {
        "label": "Amazon Web Services",
        "what": "A read-only IAM access key (not your console login).",
        "steps": [
            "AWS Console → IAM → Users → Add user.",
            "Attach the AWS-managed policy 'SecurityAudit' (read-only).",
            "Create an access key for that user.",
            "Open the secure link and paste the Access key ID, Secret access key and region.",
        ],
    },
    "azure": {
        "label": "Microsoft Azure",
        "what": "A read-only Entra ID app registration (a 'service principal') — not your Azure sign-in.",
        "steps": [
            "Microsoft Entra ID → App registrations → New registration (name: 'TrackVault read-only').",
            "Certificates & secrets → New client secret → copy the secret Value.",
            "Subscriptions → your subscription → Access control (IAM) → Add role assignment → "
            "assign BOTH 'Reader' and 'Security Reader' to that app.",
            "Open the secure link and paste the Directory (tenant) ID, Application (client) ID and Client secret.",
        ],
    },
    "intune": {
        "label": "Microsoft Intune / Defender",
        "what": "The same Entra app registration as Azure, plus one Graph permission.",
        "steps": [
            "Use the same app registration as Azure (or create one the same way).",
            "API permissions → Microsoft Graph → Application permissions → add "
            "'DeviceManagementManagedDevices.Read.All' → Grant admin consent.",
            "Open the secure link and paste the Tenant ID, Client ID and Client secret.",
        ],
    },
    "gcp": {
        "label": "Google Cloud",
        "what": "A short-lived read-only OAuth token (~1 hour).",
        "steps": [
            "As an account with Viewer / Security Reviewer, run: gcloud auth print-access-token",
            "Open the secure link and paste the Project ID and the token (do it right before, "
            "the token expires in about an hour).",
        ],
    },
    "adgpo": {
        "label": "Active Directory / GPO",
        "what": "Posture values only — no credentials, nothing connects to your directory.",
        "steps": [
            "Read your password policy, privileged-group counts and GPO settings from AD / Group Policy.",
            "Open the secure link and fill in the values (leave anything unknown blank).",
        ],
    },
    "firewall": {
        "label": "Firewall configuration",
        "what": "Config text only — no credentials.",
        "steps": [
            "Export your firewall's configuration / ruleset.",
            "Open the secure link and paste it.",
        ],
    },
}


def _hash(raw: str) -> str:
    return hashlib.sha256((raw or "").encode("utf-8")).hexdigest()


def create_request(db: Session, company: Company, providers: list[str], actor_email: str,
                   ttl_days: int = 14) -> tuple[str, AccessRequest]:
    """Create a tokenised access request. Returns (raw_token, row). Only the token
    hash is stored — the raw token lives only in the link we send."""
    raw = secrets.token_urlsafe(32)
    ar = AccessRequest(
        company_id=company.id, token_hash=_hash(raw),
        providers=[p for p in providers if p in _SCANNERS],
        created_by=actor_email,
        expires_at=datetime.now(timezone.utc) + timedelta(days=ttl_days))
    db.add(ar)
    db.commit()
    db.refresh(ar)
    return raw, ar


def resolve(db: Session, raw_token: str) -> AccessRequest | None:
    """Return the valid, unexpired, unrevoked request for this token, else None."""
    ar = db.execute(select(AccessRequest).where(
        AccessRequest.token_hash == _hash(raw_token))).scalar_one_or_none()
    if not ar or ar.status == "revoked":
        return None
    if ar.expires_at and ar.expires_at < datetime.now(timezone.utc):
        return None
    return ar


def test_connection(db: Session, company: Company, provider: str) -> tuple[bool, str]:
    """Run the connector's read-only check and return a plain-English result the
    customer and operator can trust ('read-only access confirmed…')."""
    from ..crypto import decrypt_secret
    if provider not in _SCANNERS:
        return False, "This connector can't be tested."
    conn = db.execute(select(Connector).where(
        Connector.company_id == company.id, Connector.provider == provider)).scalar_one_or_none()
    if not conn or not conn.secret_enc:
        return False, "No details saved yet for this connector."
    creds = {**(conn.public_config or {}), **decrypt_secret(conn.secret_enc), "consent": conn.consent}
    try:
        mod = __import__(_SCANNERS[provider], fromlist=["run_checks"])
        findings, _meta = mod.run_checks(creds)
    except Exception as ex:  # pragma: no cover - defensive
        return False, f"The check couldn't run ({type(ex).__name__}). Please re-check the values."
    if not findings:
        return False, "The check returned no result — please re-check the values."
    first = findings[0]
    ev = ""
    if first.get("evidence"):
        e = first["evidence"][0]
        ev = (e.get("excerpt") or e.get("note") or "").strip()
    if first.get("status") in ("ok", "partial"):
        tail = f" {ev.rstrip('.')}." if ev else ""
        return True, ("✓ Connected — read-only access confirmed." + tail +
                      " We can only read posture; we cannot change anything.")
    note = ""
    if first.get("evidence") and first["evidence"][0].get("note"):
        note = " " + first["evidence"][0]["note"]
    return False, f"✗ Couldn't connect: {ev or 'the credentials were rejected'}.{note}"


def build_request_workbook(brand: str, company_name: str, providers: list[str], link: str) -> bytes:
    """A branded instruction pack: what to create per connector, the read-only
    guarantee, and the secure link where values are entered (never in this file)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Access request"
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 100
    r = 1

    def line(text, *, size=11, bold=False, color="13202E", fill=None, wrap=True):
        nonlocal r
        cell = ws.cell(row=r, column=2, value=text)
        cell.font = Font(size=size, bold=bold, color=color)
        cell.alignment = Alignment(wrap_text=wrap, vertical="top")
        if fill:
            cell.fill = PatternFill("solid", fgColor=fill)
        r += 1

    line(f"{brand} — read-only access request", size=15, bold=True, color="147A3D")
    line(f"Prepared for: {company_name}", color="55677A")
    r += 1
    line("What this is", bold=True)
    line("We assess your compliance posture using READ-ONLY access — we can look, never "
         "change anything, and you can revoke access at any time. For each item below, your IT "
         "admin creates a read-only credential, then enters the values into the secure link. "
         "Values are never emailed.")
    r += 1
    line("① Open the secure link and keep it handy:", bold=True)
    line(link, color="1D63D8")
    r += 1
    line("② For each connector, do the following, then paste the values into the link:", bold=True)
    r += 1
    for prov in providers:
        s = SETUP_STEPS.get(prov)
        if not s:
            continue
        line(s["label"], size=12, bold=True, fill="EDF1F6")
        line(s["what"], color="55677A")
        for i, step in enumerate(s["steps"], start=1):
            line(f"   {i}. {step}")
        r += 1
    line("Questions? Reply to the email that carried this file — your engagement team will help.",
         color="55677A")

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
