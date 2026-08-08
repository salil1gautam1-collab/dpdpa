"""Generate the customer-facing Excel questionnaire template from the rulebook.

The workbook has:
  - "Start Here" sheet: purpose, what we collect, what we do with it, the
    consent statement, and how to fill it in.
  - "Questionnaire" sheet: every checkpoint grouped into sections (categories),
    each row explaining what it means, with a status dropdown that includes
    "N/A (does not apply to us)".

The same file uploads back cleanly via import_parser (header + column mapping).
"""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from ..reporting import DISCLAIMER

NAVY = "0D2137"
STEEL = "14324F"
LIGHT = "EEF2F6"
BAND = "F6F8FB"
SECTION = "1F4A6E"

_STATUS_OPTIONS = ["Compliant", "Partial", "Gap", "N/A", "To be confirmed"]

_thin = Side(style="thin", color="C9D2DB")
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _title(ws, cell, text, size=16, color=NAVY, bold=True):
    ws[cell] = text
    ws[cell].font = Font(size=size, bold=bold, color=color)


def build_template(rulebook: dict, brand: str, company_name: str = "") -> bytes:
    wb = Workbook()

    # ---------- Start Here ----------
    s = wb.active
    s.title = "Start Here"
    s.sheet_view.showGridLines = False
    s.column_dimensions["A"].width = 3
    s.column_dimensions["B"].width = 100

    _title(s, "B2", f"{brand} — DPDPA Compliance Questionnaire", size=18)
    s["B3"] = "Digital Personal Data Protection Act, 2023 & DPDP Rules, 2025"
    s["B3"].font = Font(size=11, color=STEEL)

    rows = [
        ("", ""),
        ("Company", company_name or "(type your organisation's name)"),
        ("", ""),
        ("Why you received this", None),
        ("Your DPDPA assessment team uses your answers to map your organisation against the DPDPA "
         "checkpoints and prepare your compliance report. Please complete the 'Questionnaire' tab and "
         "return this file to your engagement contact.", None),
        ("", ""),
        ("What we collect", None),
        ("Only your answers to the checkpoints: a status for each, optional evidence notes, and the "
         "responsible department. Do not paste customer personal data, passwords, or secret keys into "
         "this file — describe controls, don't include the underlying data.", None),
        ("", ""),
        ("What we do with it", None),
        ("We assess it against the DPDPA rulebook and produce your gap report with evidence and "
         "recommendations. Your data stays in your encrypted workspace and is not shared. We identify "
         "gaps; we act on your systems only where you separately grant consent, access and permission.", None),
        ("", ""),
        ("How to fill it in", None),
        ("1. Go to the 'Questionnaire' tab.  2. For each row, pick a status from the drop-down in the "
         "'Your status' column.  3. Choose 'N/A' for anything that does not apply to you (for example, an "
         "Azure section if you don't use Azure).  4. Add a short note in 'Evidence / notes' and the owning "
         "'Department'.  5. Save and return the file.", None),
        ("", ""),
        ("Status options", None),
        ("Compliant — the control is in place.   Partial — partly in place.   Gap — not in place.   "
         "N/A — does not apply to your organisation.   To be confirmed — you're not sure yet.", None),
        ("", ""),
        ("Consent", None),
        ("By completing and returning this questionnaire, the person named below confirms they are "
         "authorised to provide this information on behalf of the organisation and consent to its use for "
         "the DPDPA assessment described above.", None),
        ("", ""),
        ("Completed by (name & designation)", "(type here)"),
        ("Date", "(type here)"),
        ("I confirm the above consent (Yes / No)", "(type Yes)"),
        ("", ""),
        ("Disclaimer", None),
        (DISCLAIMER, None),
    ]
    r = 5
    for label, val in rows:
        if val is None:
            s.cell(row=r, column=2, value=label)
            cell = s.cell(row=r, column=2)
            if label in ("Why you received this", "What we collect", "What we do with it",
                         "How to fill it in", "Status options", "Consent", "Disclaimer"):
                cell.font = Font(size=12, bold=True, color=NAVY)
            else:
                cell.font = Font(size=10.5, color="333333")
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                s.row_dimensions[r].height = 42
        else:
            s.cell(row=r, column=2, value=f"{label}:  ").font = Font(bold=True, color=STEEL)
            s.cell(row=r, column=2).value = f"{label}:   {val}"
        r += 1

    # ---------- Questionnaire ----------
    q = wb.create_sheet("Questionnaire")
    q.sheet_view.showGridLines = False
    headers = ["Checkpoint ID", "Area", "Checkpoint", "What it means", "Legal basis",
               "Your status", "Evidence / notes", "Department / owner"]
    widths = [14, 20, 34, 48, 18, 16, 40, 22]
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        q.column_dimensions[get_column_letter(i)].width = w

    # intro title row
    q.merge_cells("A1:H1")
    _title(q, "A1", "DPDPA Questionnaire — pick a status for every row (choose N/A if it doesn't apply)", size=13)
    q.row_dimensions[1].height = 22

    # header row (row 3)
    hdr_row = 3
    for i, h in enumerate(headers, start=1):
        c = q.cell(row=hdr_row, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.border = _border
    q.row_dimensions[hdr_row].height = 26
    q.freeze_panes = f"A{hdr_row + 1}"

    dv = DataValidation(type="list", formula1='"' + ",".join(_STATUS_OPTIONS) + '"',
                        allow_blank=True, showDropDown=False)
    dv.prompt = "Pick one. Choose N/A if this doesn't apply to your organisation."
    dv.promptTitle = "Your status"
    q.add_data_validation(dv)

    cats = {c["id"]: c for c in rulebook["categories"]}
    # group controls by category, in category order
    order = [c["id"] for c in rulebook["categories"]]
    by_cat: dict = {}
    for ctrl in rulebook["controls"]:
        by_cat.setdefault(ctrl["category"], []).append(ctrl)

    row = hdr_row + 1
    banded = False
    for cat_id in order:
        controls = by_cat.get(cat_id)
        if not controls:
            continue
        cat = cats.get(cat_id, {"name": cat_id, "legalRef": ""})
        # section header
        q.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        sc = q.cell(row=row, column=1, value=f"{cat['name']}   —   {cat.get('legalRef','')}")
        sc.font = Font(bold=True, color="FFFFFF", size=11)
        sc.fill = PatternFill("solid", fgColor=SECTION)
        sc.alignment = Alignment(vertical="center", indent=1)
        q.row_dimensions[row].height = 22
        row += 1
        for ctrl in controls:
            vals = [ctrl["id"], cat["name"], ctrl["title"], ctrl.get("description", ""),
                    ctrl.get("legalRef", ""), "", "", ""]
            for i, v in enumerate(vals, start=1):
                c = q.cell(row=row, column=i, value=v)
                c.border = _border
                c.alignment = Alignment(wrap_text=True, vertical="top")
                if banded:
                    c.fill = PatternFill("solid", fgColor=BAND)
            dv.add(q.cell(row=row, column=6))
            q.cell(row=row, column=1).font = Font(bold=True, color=STEEL)
            q.row_dimensions[row].height = 34
            banded = not banded
            row += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
