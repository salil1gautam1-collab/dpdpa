"""Rulebook access — versioned, stored in the database."""
from __future__ import annotations

from sqlalchemy import select
from sqlalchemy.orm import Session

from ..models import Rulebook


def _version_key(v: str):
    return [int(p) if p.isdigit() else 0 for p in v.split(".")]


def all_rulebooks(db: Session) -> list[Rulebook]:
    rbs = list(db.execute(select(Rulebook)).scalars())
    return sorted(rbs, key=lambda r: _version_key(r.version))


def latest_rulebook(db: Session) -> dict:
    rbs = all_rulebooks(db)
    if not rbs:
        raise RuntimeError("No rulebook loaded")
    return rbs[-1].data


def get_rulebook(db: Session, version: str | None) -> dict:
    if not version:
        return latest_rulebook(db)
    row = db.execute(select(Rulebook).where(Rulebook.version == version)).scalar_one_or_none()
    if not row:
        raise ValueError(f"Rulebook version {version} not found")
    return row.data


def build_rulebook_workbook(rb: dict, brand: str) -> bytes:
    """The rulebook as a study document for CS/Legal: one row per control with
    every field — legal reference, how it's checked, evidence required,
    remediation — grouped by section, with a review-notes column to mark up."""
    import io
    from datetime import date
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    cats = {c["id"]: c["name"] for c in rb.get("categories", [])}
    wb = Workbook()
    ws = wb.active
    ws.title = "Rulebook"
    ws["A1"] = f"{brand} — DPDPA rulebook v{rb.get('rulebookVersion', '?')}"
    ws["A1"].font = Font(size=15, bold=True, color="147A3D")
    ws["A2"] = (f"Controls: {len(rb.get('controls', []))}   ·   Last updated: "
                f"{rb.get('lastUpdated', '?')}   ·   Exported: {date.today().isoformat()}")
    ws["A3"] = (rb.get("updateNote", "") or "")[:250]
    ws["A3"].font = Font(italic=True, size=10, color="55677A")
    ws["A4"] = ("For legal/CS review: verify each control against the DPDP Act 2023 and DPDP "
                "Rules 2025. Use the Review notes column; propose changes via Admin → Rulebook.")
    ws["A4"].font = Font(italic=True, size=10, color="8F6400")

    headers = ["Control ID", "Section", "Title", "Severity", "Legal reference", "Checked how",
               "Description", "Evidence required", "Remediation", "Review notes"]
    hrow = 6
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=hrow, column=i, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="13202E")
    for i, w in enumerate([11, 20, 34, 10, 14, 12, 48, 34, 48, 28], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{hrow + 1}"

    r = hrow
    last_cat = None
    band = PatternFill("solid", fgColor="EDF1F6")
    for c in sorted(rb.get("controls", []), key=lambda x: (x.get("category", ""), x.get("id", ""))):
        if c.get("category") != last_cat:
            r += 1
            last_cat = c.get("category")
            cell = ws.cell(row=r, column=1,
                           value=f"{last_cat} — {cats.get(last_cat, last_cat)}")
            cell.font = Font(bold=True, size=11)
            cell.fill = band
            for col in range(2, len(headers) + 1):
                ws.cell(row=r, column=col).fill = band
        r += 1
        vals = [c.get("id", ""), cats.get(c.get("category", ""), c.get("category", "")),
                c.get("title", ""), c.get("severity", ""), c.get("legalRef", ""),
                c.get("checkMethod", ""), c.get("description", ""),
                c.get("evidenceRequired", ""), c.get("remediation", ""), ""]
        for i, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=i, value=v)
            if i in (3, 7, 8, 9):
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
